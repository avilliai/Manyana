# -*- coding: utf-8 -*-
import datetime
import json
import os

import openpyxl
from openpyxl import load_workbook

#此方法用于更新词库
#导入到bot所用txt
def importDict(xlsxPath = 'data/autoReply/lexicon/public.xlsx'):
    #xlsxPath = 'D:\Mirai\YirisVoiceGUI\PythonPlugins\Config\可爱系二次元bot词库1.5万词V1.1.xlsx'


    ls1=os.listdir("data/autoReply/lexicon")
    diss={}
    for isa in ls1:
        if isa.endswith(".xlsx"):
            xlsxPath="data/autoReply/lexicon/"+isa
            # 第一步打开工作簿
            wb = openpyxl.load_workbook(xlsxPath)
            # 第二步选取表单
            sheet = wb.active
            # 按行获取数据转换成列表
            rows_data = list(sheet.rows)
            # 获取表单的表头信息(第一行)，也就是列表的第一个元素
            titles = [title.value for title in rows_data.pop(0)]
            #print(titles)


            all_row_dict = []

            #当你需要向已有词库导入时取消注释
            '''fileaa = open('Config\\superDict.txt', 'r')
            js1 = fileaa.read()
            newDict = json.loads(js1)
            print('已读取现存字典')'''

            #新建词库，当你需要新建词库时取消注释
            newDict={}

            # 遍历出除了第一行的其他行
            for a_row in rows_data:
                the_row_data = [cell.value for cell in a_row]
                # 将表头和该条数据内容，打包成一个字典
                row_dict = dict(zip(titles, the_row_data))
                #print(row_dict)
                all_row_dict.append(row_dict)
            for i in all_row_dict:
                key=i.get('key')
                value=i.get('value')

                if (key in newDict):
                    replyValue=newDict.get(key)
                    if value in replyValue:
                        pass
                        #print('已存在该回复，不添加')
                    else:
                        replyValue.append(value)
                        #print('已有关键字，追加')
                    #print(replyValue)
                # 没有关键字则创建
                else:
                    newDict[key] = [value,]
                #print('key:'+key+' '+'value:'+value)
            diss[isa.split(".")[0]]=newDict
            #print(isa+"导入完成")
    js = json.dumps(diss)

    file = open('config\superDict.txt', 'w')
    file.write(js)
    file.close()

def clearSheet(filename= "data/autoReply/lexicon/public.xlsx"):
    i=0


    while i<4:

        wb = load_workbook(filename)
        ws = wb.active
        ws.delete_cols(1)  # 删除第 1 列数据
        wb.save(filename)
        i+=1
    time1 = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    print(time1 + '| clear，已清除xlsx')
#导出到xlsx
def outPutDic():
    las=os.listdir("data/autoReply/lexicon")
    for ias in las:
        filename="data/autoReply/lexicon/"+ias
        groupId=ias.split(".")[0]
        clearSheet(filename)
        file = open('Config\\superDict.txt', 'r')
        jss = file.read()
        dict = json.loads(jss)
        dict=dict.get(groupId)
        Keys = dict.keys()
        time1 = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        print(time1 + '| 已读取字典')
        wb = openpyxl.load_workbook(filename)
        sheet = wb.active
        sheet.append(['key', 'value'])  # 插入一行数据
        for d in Keys:
            key=d
            values=dict.get(d)

            for value in values:

                #print(str(key)+str(value))

                sheet.append([key, value])  # 插入一行数据
        wb.save(filename)  # 保存,传入原文件则在
        time1 = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        print(time1 + '| 词库已同步至xlsx')



if __name__ == '__main__':
    importDict()
