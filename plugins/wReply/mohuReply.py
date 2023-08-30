# -*- coding: utf-8 -*-
import datetime
import os
import random
import sys

import json


#可以优化，bot.py中增加方式已经更新，但我懒得改
#import openpyxl
#from fuzzywuzzy import fuzz
import openpyxl


def mohuaddReplys(ass,groupid,mode=0):
    message=ass[2:]
    messageS=message.split('#')
    file = open('config/superDict.txt', 'r')
    js = file.read()
    dict = json.loads(js)
    file.close()
    print(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S') + '| 已读取字典')
    if str(groupid) not in dict:
        print(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S') + '| 创建新文件' + groupid + ".xlsx")
        dict1 = {messageS[0]: [messageS[1]]}

        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(["key", "value"])
    else:
        dict1 = dict.get(groupid)
    #对传入的字符串进行处理并加入字典
    #如果已经有关键字
    if (messageS[0] in dict1):
        replyValue = dict1.get(messageS[0])
        replyValue.append(messageS[1])
        print(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')+ '| 已有关键字，追加')
        # print(replyValue)
    # 没有关键字则创建
    else:
        dict1[messageS[0]] = [messageS[1], ]
    if mode==1:
        wb = openpyxl.load_workbook("data/autoReply/lexicon/" +"publicLexicon.xlsx")
        sheet = wb.active
        sheet.append([messageS[0], messageS[1]])
        wb.save("data/autoReply/lexicon/" +"publicLexicon.xlsx")
    else:
        wb = openpyxl.load_workbook("data/autoReply/lexicon/" + groupid + ".xlsx")
        sheet = wb.active
            #print(dict)
        #重新写入
        #print(dict1)
        #print(dict)

        #写入excel
        sheet.append([messageS[0], messageS[1]])  # 插入一行数据
        wb.save("data/autoReply/lexicon/"+groupid+".xlsx")  # 保存,传入原文件则在

    dict[str(groupid)] = dict1

    #print(dict)
    #print(type(dict))
    js = json.dumps(dict)
    file = open('config/superDict.txt', 'w')
    file.write(js)
    file.close()
    #print(type(dict))
    return dict
def mohuadd(key,value,group):
    file = open('config/superDict.txt', 'r')
    js = file.read()
    dict = json.loads(js)
    print(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')+ '| 已读取字典')
    #print(dict)
    #print('---------')
    file.close()
    dag=dict.get(group)
    dag[key]=value
    dict[group]=dag
        #print(dict)
    #重新写入
    #print(dict)
    js = json.dumps(dict)
    file = open('config/superDict.txt', 'w')
    file.write(js)
    file.close()
    #print(type(dict))
    return dict






def mohudels(messagess,group2):
    file = open('config/superDict.txt', 'r')
    js = file.read()
    dict = json.loads(js)
    dicaa=dict.get(str(group2))
    dicaa.pop(messagess)
    dict[str(group2)]=dicaa



    js = json.dumps(dict)
    file = open('config/superDict.txt', 'w')
    file.write(js)
    file.close()
    return dict

def mohudelValue(key,valueNo):
    file = open('config/superDict.txt', 'r')
    js = file.read()
    dict = json.loads(js)

    if key in dict.keys():
        values = dict.get(key)
        try:
            value1 = values.remove(valueNo)
            dict[key] = value1
        except:
            print('error')
    else:
        print(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')+ '| 没有指定词')
    js = json.dumps(dict)
    file = open('Config/superDict.txt', 'w')
    file.write(js)
    file.close()
    return dict



#with open("config\\replyDic.txt",'a') as f:
if __name__ == '__main__':
    '''print('当前路径' + sys.argv[0])
    file = open('Config\\dict.txt', 'r')
    js = file.read()
    dict = json.loads(js)
    print('已读取字典')
    print(dict)
    while True:
        s=input('输入命令')
        if s.startswith('添加'):
            print(addReplys(s))
        elif s.startswith('删除'):
            print(dels(s))'''
    mohudels('模糊删除图片')

